import os
from typing import Dict, Any
from openpyxl import load_workbook
from backend.converters.base import BaseConverter

class ExcelToHtmlConverter(BaseConverter):
    """
    Excel to HTML Converter
    Converts Excel files (.xlsx, .xls) to HTML format
    """
    
    def convert(self, input_path: str, output_path: str, **options) -> Dict[str, Any]:
        self.validate_input(input_path, 5)
        
        try:
            # Load workbook
            wb = load_workbook(input_path, data_only=True)
            self.update_progress(input_path, 20)
            
            html_content = []
            html_content.append('<!DOCTYPE html>')
            html_content.append('<html><head><meta charset="utf-8"><style>')
            html_content.append('table { border-collapse: collapse; width: 100%; margin-bottom: 20px; }')
            html_content.append('th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }')
            html_content.append('th { background-color: #f2f2f2; }')
            html_content.append('.sheet-title { font-size: 1.5em; margin: 20px 0 10px; font-weight: bold; }')
            html_content.append('</style></head><body>')
            
            total_sheets = len(wb.sheetnames)
            for index, sheet_name in enumerate(wb.sheetnames):
                ws = wb[sheet_name]
                html_content.append(f'<div class="sheet-title">{sheet_name}</div>')
                html_content.append('<table>')
                
                rows = list(ws.iter_rows(values_only=True))
                if rows:
                    # Header
                    html_content.append('<thead><tr>')
                    for cell in rows[0]:
                        html_content.append(f'<th>{cell if cell is not None else ""}</th>')
                    html_content.append('</tr></thead>')
                    
                    # Body
                    html_content.append('<tbody>')
                    for row in rows[1:]:
                        html_content.append('<tr>')
                        for cell in row:
                            html_content.append(f'<td>{cell if cell is not None else ""}</td>')
                        html_content.append('</tr>')
                    html_content.append('</tbody>')
                
                html_content.append('</table>')
                
                # Update progress
                progress = 20 + int(((index + 1) / total_sheets) * 70)
                self.update_progress(input_path, progress)
            
            html_content.append('</body></html>')
            
            # Write to file
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write('\n'.join(html_content))
                
            self.update_progress(input_path, 100)
            return {'success': True, 'output_path': output_path}
            
        except Exception as e:
            return {'success': False, 'error': str(e)}
