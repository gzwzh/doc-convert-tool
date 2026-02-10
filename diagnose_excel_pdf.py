#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
诊断Excel转PDF问题
"""

import os
import sys

# 添加backend到路径
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'backend'))

print("=" * 60)
print("Excel转PDF功能诊断")
print("=" * 60)

# 1. 检查转换器代码
print("\n1. 检查转换器代码...")
try:
    from converters.excel_to_pdf import ExcelToPdfConverter
    converter = ExcelToPdfConverter()
    
    # 检查是否有优化方法
    has_convert_with_excel = hasattr(converter, '_convert_with_excel')
    print(f"   ✓ ExcelToPdfConverter 已导入")
    print(f"   ✓ _convert_with_excel 方法存在: {has_convert_with_excel}")
    
    # 检查代码内容
    import inspect
    source = inspect.getsource(converter._convert_with_excel)
    
    # 检查关键优化点
    checks = {
        "智能纸张选择": "col_count <= 6" in source,
        "列宽余量": "col.ColumnWidth = curr_width * 1.15" in source or "col.ColumnWidth * 1.15" in source,
        "缩放策略": "setup.Zoom = 85" in source,
        "高质量导出": "Quality=0" in source or "PrintQuality = 600" in source,
    }
    
    print("\n   优化特性检查:")
    for feature, exists in checks.items():
        status = "✓" if exists else "✗"
        print(f"   {status} {feature}: {'已实现' if exists else '未找到'}")
    
    all_ok = all(checks.values())
    if all_ok:
        print(f"\n   ✅ 所有优化特性都已实现")
    else:
        print(f"\n   ⚠️ 部分优化特性未找到，可能代码未更新")
        
except Exception as e:
    print(f"   ❌ 错误: {e}")
    import traceback
    traceback.print_exc()

# 2. 检查服务注册
print("\n2. 检查服务注册...")
try:
    from services.converter_service import ConverterService
    service = ConverterService()
    
    # 检查Excel转PDF是否注册
    excel_pdf_registered = ('xlsx', 'pdf') in service.converters
    xls_pdf_registered = ('xls', 'pdf') in service.converters
    
    print(f"   ✓ ConverterService 已导入")
    print(f"   ✓ xlsx->pdf 已注册: {excel_pdf_registered}")
    print(f"   ✓ xls->pdf 已注册: {xls_pdf_registered}")
    
    if excel_pdf_registered:
        converter_instance = service.converters[('xlsx', 'pdf')]
        converter_class = type(converter_instance).__name__
        print(f"   ✓ 使用的转换器: {converter_class}")
        
        if converter_class == "ExcelToPdfConverter":
            print(f"   ✅ 正在使用重构后的转换器")
        else:
            print(f"   ⚠️ 使用的不是ExcelToPdfConverter")
    
except Exception as e:
    print(f"   ❌ 错误: {e}")
    import traceback
    traceback.print_exc()

# 3. 检查依赖
print("\n3. 检查依赖...")
try:
    import comtypes
    print(f"   ✓ comtypes 已安装 (版本: {comtypes.__version__ if hasattr(comtypes, '__version__') else '未知'})")
except ImportError:
    print(f"   ✗ comtypes 未安装")
    print(f"   提示: 运行 'pip install comtypes' 安装")

try:
    import openpyxl
    print(f"   ✓ openpyxl 已安装 (版本: {openpyxl.__version__})")
except ImportError:
    print(f"   ✗ openpyxl 未安装")

# 4. 检查Excel COM可用性
print("\n4. 检查Excel COM可用性...")
if sys.platform == 'win32':
    try:
        import comtypes.client
        excel = comtypes.client.CreateObject('Excel.Application')
        excel.Quit()
        print(f"   ✓ Microsoft Excel COM 可用")
    except Exception as e:
        print(f"   ✗ Microsoft Excel COM 不可用: {e}")
        print(f"   提示: 确保已安装Microsoft Excel")
else:
    print(f"   ⚠️ 非Windows系统，Excel COM不可用")

# 5. 测试转换
print("\n5. 快速转换测试...")
try:
    import openpyxl
    from pathlib import Path
    
    # 创建测试文件
    wb = openpyxl.Workbook()
    ws = wb.active
    for col in range(1, 21):
        ws.cell(row=1, column=col, value=f"列{col}")
    for row in range(2, 6):
        for col in range(1, 21):
            ws.cell(row=row, column=col, value=f"数据{row}-{col}")
    
    test_dir = Path("backend/downloads")
    test_dir.mkdir(parents=True, exist_ok=True)
    test_excel = test_dir / "diagnose_test.xlsx"
    wb.save(test_excel)
    
    print(f"   ✓ 创建测试Excel: {test_excel}")
    print(f"   ✓ 列数: 20 (应使用A3横向，85%缩放)")
    
    # 转换
    from converters.excel_to_pdf import ExcelToPdfConverter
    converter = ExcelToPdfConverter()
    test_pdf = test_dir / "diagnose_test.pdf"
    
    result = converter.convert(str(test_excel), str(test_pdf))
    
    if result.get('success'):
        print(f"   ✅ 转换成功")
        print(f"   ✓ 方法: {result.get('method')}")
        print(f"   ✓ 输出: {test_pdf}")
        print(f"   ✓ 大小: {result.get('size', 0) / 1024:.2f} KB")
        print(f"\n   请打开 {test_pdf} 检查效果")
    else:
        print(f"   ❌ 转换失败: {result.get('error')}")
        
except Exception as e:
    print(f"   ❌ 测试失败: {e}")
    import traceback
    traceback.print_exc()

# 总结
print("\n" + "=" * 60)
print("诊断总结")
print("=" * 60)
print("\n如果所有检查都通过，但转换效果仍不理想，请检查:")
print("1. 应用是否已重启（Python后端需要重启才能加载新代码）")
print("2. 浏览器缓存是否已清除")
print("3. 原始Excel文件是否有特殊格式或内容")
print("4. 查看生成的测试PDF文件效果")
print("\n如果测试PDF效果正常，说明代码没问题，可能是:")
print("- 应用未重启")
print("- 使用了缓存的旧版本")
print("- 特定Excel文件的问题")
