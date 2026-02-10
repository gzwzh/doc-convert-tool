#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
分析你的Excel文件
"""

import openpyxl
from pathlib import Path
import sys

print("=" * 60)
print("分析Excel文件")
print("=" * 60)

# 查找最近的Excel文件
downloads_dir = Path("backend/downloads")
if not downloads_dir.exists():
    print("downloads目录不存在")
    exit(1)

excel_files = list(downloads_dir.glob("*.xlsx")) + list(downloads_dir.glob("*.xls"))
if not excel_files:
    print("未找到Excel文件")
    exit(1)

# 按修改时间排序，获取最新的
excel_files.sort(key=lambda x: x.stat().st_mtime, reverse=True)
latest_file = excel_files[0]

print(f"\n分析文件: {latest_file.name}")
print(f"文件大小: {latest_file.stat().st_size / 1024:.2f} KB")

try:
    wb = openpyxl.load_workbook(latest_file, data_only=True)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # 获取实际使用的区域
        max_row = ws.max_row
        max_col = ws.max_column
        
        print(f"\n工作表: {sheet_name}")
        print(f"  最大行数: {max_row}")
        print(f"  最大列数: {max_col}")
        
        # 根据列数给出建议
        if max_col <= 6:
            print(f"  📄 建议: A4纵向")
            print(f"  ✓ 这是窄表格，应该效果很好")
        elif max_col <= 10:
            print(f"  📄 建议: A4横向")
            print(f"  ✓ 中等宽度，应该效果不错")
        elif max_col <= 15:
            print(f"  📄 建议: A3横向，适应1页宽")
            print(f"  ✓ 较宽表格，会自动调整")
        else:
            print(f"  📄 建议: A3横向，85%固定缩放")
            print(f"  ⚠️ 超宽表格，可能需要手动调整列宽")
        
        # 显示前几行数据
        print(f"\n  前3行数据预览:")
        for row_idx in range(1, min(4, max_row + 1)):
            row_data = []
            for col_idx in range(1, min(max_col + 1, 11)):  # 最多显示10列
                cell = ws.cell(row=row_idx, column=col_idx)
                value = str(cell.value) if cell.value else ""
                if len(value) > 15:
                    value = value[:12] + "..."
                row_data.append(value)
            
            if max_col > 10:
                row_data.append(f"... (共{max_col}列)")
            
            print(f"    行{row_idx}: {' | '.join(row_data)}")
    
    wb.close()
    
    print(f"\n" + "=" * 60)
    print("分析完成")
    print("=" * 60)
    
    if max_col > 15:
        print(f"\n💡 建议:")
        print(f"  你的表格有 {max_col} 列，属于超宽表格")
        print(f"  优化后会使用85%固定缩放，保证清晰度")
        print(f"  如果还是觉得小，可以:")
        print(f"  1. 在Excel中先调整列宽")
        print(f"  2. 或者拆分为多个工作表")
        print(f"  3. 或者设置打印区域")
    
except Exception as e:
    print(f"\n❌ 分析失败: {e}")
    import traceback
    traceback.print_exc()
