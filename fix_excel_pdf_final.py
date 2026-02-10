#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
最终修复Excel转PDF问题
"""

import sys
import os

# 添加backend到路径
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'backend'))

print("=" * 60)
print("最终修复Excel转PDF")
print("=" * 60)

# 1. 验证代码
print("\n1. 验证优化代码...")
from converters.excel_to_pdf import ExcelToPdfConverter
import inspect

converter = ExcelToPdfConverter()
source = inspect.getsource(converter._convert_with_excel)

checks = {
    "智能纸张选择": "col_count <= 6" in source and "setup.PaperSize = 8" in source,
    "列宽余量": "1.15" in source,
    "缩放策略": "setup.Zoom = 85" in source,
    "A3横向": "setup.PaperSize = 8" in source and "setup.Orientation = 2" in source,
}

all_ok = True
for feature, exists in checks.items():
    status = "✅" if exists else "❌"
    print(f"  {status} {feature}")
    if not exists:
        all_ok = False

if not all_ok:
    print("\n❌ 代码验证失败！请检查backend/converters/excel_to_pdf.py")
    exit(1)

print("\n✅ 代码验证通过")

# 2. 创建测试Excel
print("\n2. 创建测试Excel...")
import openpyxl
from pathlib import Path

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "测试"

# 创建20列数据
for col in range(1, 21):
    ws.cell(row=1, column=col, value=f"列{col}")

for row in range(2, 6):
    for col in range(1, 21):
        ws.cell(row=row, column=col, value=f"数据{row}-{col}")

test_dir = Path("backend/downloads")
test_dir.mkdir(parents=True, exist_ok=True)
test_file = test_dir / "final_test_20cols.xlsx"
wb.save(test_file)

print(f"  ✅ 创建: {test_file}")
print(f"  ✓ 20列数据")

# 3. 直接转换
print("\n3. 直接转换（不通过API）...")

def progress_callback(file_path, progress):
    if progress % 25 == 0 or progress == 100:
        print(f"  进度: {progress}%")

converter.set_progress_callback(progress_callback)

output_file = test_dir / "final_test_20cols.pdf"
result = converter.convert(str(test_file), str(output_file))

if result.get('success'):
    print(f"\n  ✅ 转换成功")
    print(f"  ✓ 方法: {result.get('method')}")
    print(f"  ✓ 输出: {output_file}")
    
    # 验证PDF
    import PyPDF2
    with open(output_file, 'rb') as f:
        reader = PyPDF2.PdfReader(f)
        page = reader.pages[0]
        width = float(page.mediabox.width)
        height = float(page.mediabox.height)
        width_mm = width * 0.352778
        height_mm = height * 0.352778
        
        print(f"\n  PDF信息:")
        print(f"  ✓ 页数: {len(reader.pages)}")
        print(f"  ✓ 尺寸: {width_mm:.1f}mm x {height_mm:.1f}mm")
        
        # 判断纸张
        if abs(width_mm - 420) < 10 and abs(height_mm - 297) < 10:
            print(f"  ✅ 纸张: A3横向（正确！）")
            print(f"\n  🎉 优化代码工作正常！")
        elif abs(width_mm - 210) < 10 and abs(height_mm - 297) < 10:
            print(f"  ❌ 纸张: A4纵向（错误！）")
            print(f"\n  ⚠️ 这说明使用了旧代码")
        else:
            print(f"  ⚠️ 纸张: 未知")
else:
    print(f"\n  ❌ 转换失败: {result.get('error')}")
    exit(1)

# 4. 总结
print("\n" + "=" * 60)
print("诊断结论")
print("=" * 60)

print("\n如果上面显示'A3横向（正确！）'，说明:")
print("  ✅ 优化代码本身没问题")
print("  ✅ 直接调用转换器工作正常")
print("  ❌ 但Electron应用可能使用了旧代码或缓存")

print("\n解决方案:")
print("  1. 完全关闭Electron应用")
print("  2. 删除可能的缓存:")
print("     - backend/__pycache__")
print("     - backend/converters/__pycache__")
print("  3. 重新启动应用")

print("\n或者，重新打包应用:")
print("  npm run build")

print(f"\n请打开以下文件验证效果:")
print(f"  {output_file}")
print(f"\n如果这个PDF显示正常（文字清晰，A3横向），")
print(f"说明代码没问题，只是应用需要重启或重新打包。")
