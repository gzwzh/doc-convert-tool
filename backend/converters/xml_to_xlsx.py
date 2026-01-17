import xml.etree.ElementTree as ET
from .base import BaseConverter
from typing import Dict, Any, List


class XmlToXlsxConverter(BaseConverter):
    """XML 到 XLSX 转换器"""
    
    def __init__(self):
        super().__init__()
        self.supported_formats = ['xlsx']
    
    def _extract_records(self, root) -> tuple:
        """从 XML 中提取记录数据"""
        records = []
        headers = set()
        
        # 尝试找到重复的子元素作为记录
        children = list(root)
        if not children:
            return [], []
        
        # 检查是否所有子元素标签相同
        tags = [child.tag for child in children]
        most_common_tag = max(set(tags), key=tags.count)
        
        for child in children:
            if child.tag == most_common_tag:
                record = {}
                # 提取属性
                for key, value in child.attrib.items():
                    record[key] = value
                    headers.add(key)
                
                # 提取子元素
                for sub in child:
                    tag = sub.tag.split('}')[-1] if '}' in sub.tag else sub.tag
                    record[tag] = sub.text or ''
                    headers.add(tag)
                
                # 如果没有子元素，使用文本内容
                if not list(child) and child.text:
                    record['value'] = child.text.strip()
                    headers.add('value')
                
                records.append(record)
        
        return list(headers), records
    
    def convert(self, input_path: str, output_path: str, **options) -> Dict[str, Any]:
        """将 XML 转换为 XLSX"""
        try:
            self.validate_input(input_path)
            
            # 尝试导入 openpyxl
            try:
                from openpyxl import Workbook
            except ImportError:
                raise Exception("需要安装 openpyxl: pip install openpyxl")
            
            tree = ET.parse(input_path)
            root = tree.getroot()
            
            headers, records = self._extract_records(root)
            
            if not records:
                raise Exception("无法从 XML 中提取表格数据")
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            
            # 写入表头
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # 写入数据
            for row_idx, record in enumerate(records, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=record.get(header, ''))
            
            wb.save(output_path)
            
            return {
                'success': True,
                'output_path': output_path,
                'size': self.get_output_size(output_path),
                'rows': len(records)
            }
            
        except Exception as e:
            self.cleanup_on_error(output_path)
            raise Exception(f"XML to XLSX conversion failed: {str(e)}")
