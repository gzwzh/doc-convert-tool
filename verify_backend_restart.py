#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
验证后端重启后的Excel转PDF功能
"""

import requests
import os
from pathlib import Path
import openpyxl

print("=" * 60)
print("验证后端重启后的Excel转PDF功能")
print("=" * 60)

# 1. 检查后端健康状态
print("\n1. 检查后端服务...")
try:
    response = requests.get("http://localhost:8002/api/health", timeout=5)
    if response.status_code == 200:
        print("   ✅ 后端服务正常运行")
        data = response.json()
        print(f"   ✓ 状态: {data.get('status')}")
    else:
        print(f"   ❌ 后端服务异常: {response.status_code}")
        exit(1)
except Exception as e:
    print(f"   ❌ 无法连接到后端: {e}")
    print("   提示: 确保后端服务在 http://localhost:8002 运行")
    exit(1)

# 2. 创建测试Excel文件
print("\n2. 创建测试Excel文件...")
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "测试数据"

# 创建20列的宽表格
for col in range(1, 21):
    ws.cell(row=1, column=col, value=f"列{col}")

for row in range(2, 12):
    for col in range(1, 21):
        ws.cell(row=row, column=col, value=f"数据{row}-{col}")

test_dir = Path("backend/downloads")
test_dir.mkdir(parents=True, exist_ok=True)
test_file = test_dir / "api_test_20cols.xlsx"
wb.save(test_file)

print(f"   ✅ 创建测试文件: {test_file}")
print(f"   ✓ 列数: 20 (应使用A3横向，85%固定缩放)")

# 3. 通过API转换
print("\n3. 通过API转换Excel到PDF...")
try:
    with open(test_file, 'rb') as f:
        files = {'file': ('api_test_20cols.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        data = {'target_format': 'pdf'}
        
        response = requests.post(
            "http://localhost:8002/api/convert/general",
            files=files,
            data=data,
            timeout=30
        )
    
    if response.status_code == 200:
        result = response.json()
        print(f"   ✅ 转换成功")
        print(f"   ✓ 输出文件: {result.get('filename')}")
        print(f"   ✓ 下载URL: {result.get('download_url')}")
        
        # 下载PDF文件
        if result.get('download_url'):
            pdf_url = f"http://localhost:8002{result.get('download_url')}"
            pdf_response = requests.get(pdf_url, timeout=10)
            
            if pdf_response.status_code == 200:
                output_pdf = test_dir / "api_test_20cols_result.pdf"
                with open(output_pdf, 'wb') as f:
                    f.write(pdf_response.content)
                
                pdf_size = len(pdf_response.content)
                print(f"   ✓ PDF已下载: {output_pdf}")
                print(f"   ✓ 文件大小: {pdf_size / 1024:.2f} KB")
                
                if pdf_size > 1000:  # 至少1KB
                    print(f"\n   🎉 转换成功！请打开以下文件查看效果:")
                    print(f"      {output_pdf}")
                    print(f"\n   预期效果:")
                    print(f"   - 纸张: A3横向")
                    print(f"   - 缩放: 85%固定缩放")
                    print(f"   - 文字: 清晰可读，不会缩成一团")
                    print(f"   - 列宽: 自动调整+15%余量")
                else:
                    print(f"   ⚠️ PDF文件太小，可能转换失败")
            else:
                print(f"   ❌ 下载PDF失败: {pdf_response.status_code}")
    else:
        print(f"   ❌ 转换失败: {response.status_code}")
        print(f"   错误: {response.text}")
        
except Exception as e:
    print(f"   ❌ API调用失败: {e}")
    import traceback
    traceback.print_exc()

# 4. 总结
print("\n" + "=" * 60)
print("验证总结")
print("=" * 60)
print("\n如果转换成功，说明:")
print("✅ 后端服务已重启")
print("✅ 新的Excel转PDF代码已生效")
print("✅ 优化特性已应用")
print("\n现在可以通过Web界面测试Excel转PDF功能了！")
print("应该看到文字清晰，内容完整，不会缩成一团。")
