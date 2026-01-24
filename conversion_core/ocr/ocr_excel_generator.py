#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
OCR Excel生成器 - 基于OCR结果生成Excel文档
支持: Markdown表格、HTML表格、表格去重
"""

import os
import re
import hashlib
from typing import Dict, List, Optional, Any, Set
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from conversion_core.services.markdown_parser import MarkdownParser, TableData, HTMLTableParser


@dataclass
class OCRPageResult:
    """OCR页面结果模型"""
    page_number: int
    markdown_text: str
    tables: List[Dict]
    images: List[Dict]
    success: bool
    error: Optional[str] = None


class OCRExcelGenerator:
    """基于OCR结果生成Excel文档"""
    
    HEADER_FILL = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    HEADER_FONT = Font(bold=True)
    THIN_BORDER = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # HTML表格模式
    HTML_TABLE_PATTERN = re.compile(r'<table[^>]*>.*?</table>', re.DOTALL | re.IGNORECASE)
    
    def __init__(self, output_dir: str, config: Dict = None):
        self.output_dir = output_dir
        self.config = config or {}
        self.mode = self.config.get("mode", "per_page")
        self.parser = MarkdownParser()
        self._table_hashes: Set[str] = set()  # 用于表格去重
        
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)
    
    def _get_table_hash(self, headers: List[str], rows: List[List[str]]) -> str:
        """计算表格内容的哈希值，用于去重"""
        content = str(headers) + str(rows)
        return hashlib.md5(content.encode()).hexdigest()

    def generate(self, ocr_results: List[Any], output_filename: str) -> Dict:
        """生成Excel文档"""
        try:
            wb = Workbook()
            default_sheet = wb.active
            self._table_hashes.clear()  # 重置表格哈希
            
            all_tables = []
            page_tables = {}
            
            for page_result in ocr_results:
                if not self._is_successful_page(page_result):
                    continue
                
                page_num = self._get_page_number(page_result)
                markdown_text = self._get_markdown_text(page_result)
                
                # 直接从原始文本提取表格
                tables = self._extract_all_tables(markdown_text)
                
                if tables:
                    page_tables[page_num] = tables
                    all_tables.extend(tables)
            
            # 根据模式创建工作表
            if self.mode == "per_page":
                self._create_per_page_worksheets(wb, page_tables)
            else:
                self._create_merged_worksheet(wb, all_tables)
            
            # 如果没有表格，创建空工作表
            if len(wb.worksheets) == 0:
                ws = wb.create_sheet("Sheet1")
                ws['A1'] = "未识别到表格数据"
            
            # 删除默认工作表
            if default_sheet in wb.worksheets and len(wb.worksheets) > 1:
                wb.remove(default_sheet)
            
            if not output_filename.lower().endswith('.xlsx'):
                output_filename += '.xlsx'
            
            output_path = os.path.join(self.output_dir, output_filename) if self.output_dir else output_filename
            wb.save(output_path)
            
            return {
                "success": True,
                "output_file": output_path,
                "error": None,
                "table_count": len(all_tables),
                "sheet_count": len(wb.worksheets)
            }
            
        except Exception as e:
            return {
                "success": False,
                "output_file": None,
                "error": f"生成Excel文档失败: {str(e)}"
            }
    
    def _is_successful_page(self, page_result: Any) -> bool:
        if isinstance(page_result, dict):
            return page_result.get("success", False)
        return getattr(page_result, 'success', False)
    
    def _get_page_number(self, page_result: Any) -> int:
        if isinstance(page_result, dict):
            return page_result.get("page_number", 1)
        return getattr(page_result, 'page_number', 1)
    
    def _get_markdown_text(self, page_result: Any) -> str:
        if isinstance(page_result, dict):
            return page_result.get("markdown_text", "")
        return getattr(page_result, 'markdown_text', "")

    def _extract_all_tables(self, text: str) -> List[TableData]:
        """从文本中提取所有表格（HTML和Markdown），自动去重，专注于表格数据"""
        tables = []
        
        if not text:
            return tables
        
        # 1. 提取HTML表格
        for match in self.HTML_TABLE_PATTERN.finditer(text):
            html_table = match.group(0)
            try:
                parser = HTMLTableParser()
                parser.feed(html_table)
                
                for table_dict in parser.get_tables():
                    headers = table_dict.get('headers', [])
                    rows = table_dict.get('rows', [])
                    
                    # 如果没有表头但有数据，第一行作为表头
                    if not headers and rows:
                        headers = rows[0]
                        rows = rows[1:] if len(rows) > 1 else []
                    
                    # 过滤无效表格
                    if not self._is_valid_table(headers, rows):
                        continue
                    
                    # 检查是否重复
                    table_hash = self._get_table_hash(headers, rows)
                    if table_hash in self._table_hashes:
                        print(f"[Excel跳过重复HTML表格] hash={table_hash[:8]}")
                        continue
                    self._table_hashes.add(table_hash)
                    
                    all_rows = ([headers] if headers else []) + rows
                    col_count = max(len(r) for r in all_rows) if all_rows else 0
                    
                    tables.append(TableData(
                        rows=rows,
                        headers=headers,
                        row_count=len(rows),
                        col_count=col_count,
                        source="html"
                    ))
                    print(f"[Excel添加HTML表格] 表头:{len(headers)}, 数据行:{len(rows)}, 列数:{col_count}")
                    
            except Exception as e:
                print(f"[Excel HTML表格解析失败] {e}")
                continue
        
        # 2. 移除HTML后提取Markdown表格
        text_no_html = self.HTML_TABLE_PATTERN.sub('', text)
        # 移除HTML块标签
        text_no_html = re.sub(r'<(?:div|html|body|center)[^>]*>', '', text_no_html, flags=re.IGNORECASE)
        text_no_html = re.sub(r'</(?:div|html|body|center)>', '', text_no_html, flags=re.IGNORECASE)
        # 移除其他HTML标签
        text_no_html = re.sub(r'<[^>]+>', '', text_no_html)
        
        md_tables = self._extract_markdown_tables(text_no_html)
        
        # 对Markdown表格也进行去重和验证
        for table in md_tables:
            if not self._is_valid_table(table.headers, table.rows):
                continue
                
            table_hash = self._get_table_hash(table.headers, table.rows)
            if table_hash in self._table_hashes:
                print(f"[Excel跳过重复Markdown表格] hash={table_hash[:8]}")
                continue
            self._table_hashes.add(table_hash)
            tables.append(table)
            print(f"[Excel添加Markdown表格] 表头:{len(table.headers)}, 数据行:{len(table.rows)}, 列数:{table.col_count}")
        
        return tables
    
    def _is_valid_table(self, headers: List[str], rows: List[List[str]]) -> bool:
        """验证表格是否有效"""
        # 至少要有数据行或表头
        if not rows and not headers:
            return False
        
        # 如果只有表头，至少要有2列
        if not rows and len(headers) < 2:
            return False
        
        # 如果有数据行，至少要有1行数据
        if rows and len(rows) < 1:
            return False
        
        # 检查是否有实际内容（不全是空值）
        all_cells = []
        if headers:
            all_cells.extend(headers)
        for row in rows:
            all_cells.extend(row)
        
        non_empty_cells = [cell for cell in all_cells if cell and str(cell).strip()]
        
        # 至少要有3个非空单元格才认为是有效表格
        return len(non_empty_cells) >= 3
    
    def _extract_markdown_tables(self, text: str) -> List[TableData]:
        """提取Markdown表格"""
        tables = []
        lines = text.split('\n')
        
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            
            if line.startswith('|') and line.endswith('|'):
                table_lines = [line]
                i += 1
                
                while i < len(lines):
                    next_line = lines[i].strip()
                    if next_line.startswith('|') and next_line.endswith('|'):
                        table_lines.append(next_line)
                        i += 1
                    elif not next_line:
                        i += 1
                        break
                    else:
                        break
                
                table = self._parse_md_table(table_lines)
                if table and (table.row_count > 0 or table.headers):
                    tables.append(table)
            else:
                i += 1
        
        return tables
    
    def _parse_md_table(self, lines: List[str]) -> Optional[TableData]:
        """解析Markdown表格"""
        if not lines:
            return None
        
        rows = []
        headers = []
        separator_found = False
        
        for idx, line in enumerate(lines):
            line = line.strip()
            if line.startswith('|'):
                line = line[1:]
            if line.endswith('|'):
                line = line[:-1]
            
            cells = [c.strip() for c in line.split('|')]
            
            if all(self._is_separator(c) for c in cells):
                separator_found = True
                continue
            
            if not separator_found and idx == 0:
                headers = cells
            else:
                rows.append(cells)
        
        if not separator_found and headers:
            rows.insert(0, headers)
            headers = []
        
        all_rows = ([headers] if headers else []) + rows
        col_count = max(len(r) for r in all_rows) if all_rows else 0
        
        return TableData(
            rows=rows,
            headers=headers,
            row_count=len(rows),
            col_count=col_count,
            source="markdown"
        )
    
    def _is_separator(self, cell: str) -> bool:
        cell = cell.strip()
        if not cell:
            return True
        return all(c in '-: ' for c in cell) and '-' in cell
    
    def _create_per_page_worksheets(self, wb: Workbook, page_tables: Dict[int, List[TableData]]):
        """每页创建单独的工作表"""
        for page_num in sorted(page_tables.keys()):
            tables = page_tables[page_num]
            
            sheet_name = f"第{page_num}页"
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]
            
            ws = wb.create_sheet(title=sheet_name)
            
            current_row = 1
            for idx, table in enumerate(tables):
                if idx > 0:
                    current_row += 2
                current_row = self._write_table_to_worksheet(ws, table, current_row)
    
    def _create_merged_worksheet(self, wb: Workbook, all_tables: List[TableData]):
        """创建合并的工作表"""
        ws = wb.create_sheet(title="合并数据")
        
        current_row = 1
        for idx, table in enumerate(all_tables):
            if idx > 0:
                current_row += 3
            current_row = self._write_table_to_worksheet(ws, table, current_row)

    def _write_table_to_worksheet(self, ws, table: TableData, start_row: int) -> int:
        """将表格写入工作表"""
        if isinstance(table, TableData):
            headers = table.headers
            rows = table.rows
        elif isinstance(table, dict):
            headers = table.get("headers", [])
            rows = table.get("rows", [])
        else:
            return start_row
        
        current_row = start_row
        
        # 写入表头
        if headers:
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=str(header) if header else "")
                cell.font = self.HEADER_FONT
                cell.fill = self.HEADER_FILL
                cell.border = self.THIN_BORDER
                cell.alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
        
        # 写入数据行
        for row_data in rows:
            for col_idx, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=str(cell_value) if cell_value else "")
                cell.border = self.THIN_BORDER
                cell.alignment = Alignment(vertical='center', wrap_text=True)
            current_row += 1
        
        # 自动调整列宽
        self._auto_adjust_column_width(ws, headers, rows)
        
        return current_row
    
    def _auto_adjust_column_width(self, ws, headers: List[str], rows: List[List[str]]):
        """自动调整列宽"""
        all_rows = ([headers] if headers else []) + rows
        
        if not all_rows:
            return
        
        col_count = max(len(row) for row in all_rows) if all_rows else 0
        
        for col_idx in range(col_count):
            max_length = 0
            
            for row in all_rows:
                if col_idx < len(row):
                    cell_value = str(row[col_idx]) if row[col_idx] else ""
                    length = sum(2 if ord(c) > 127 else 1 for c in cell_value)
                    max_length = max(max_length, length)
            
            adjusted_width = min(max(max_length + 2, 8), 50)
            column_letter = get_column_letter(col_idx + 1)
            ws.column_dimensions[column_letter].width = adjusted_width
